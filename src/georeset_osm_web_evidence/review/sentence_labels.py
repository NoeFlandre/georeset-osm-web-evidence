from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

SENTENCE_REVIEW_COLUMNS = [
    "review_id",
    "human_label",
    "human_notes",
    "llm_label",
    "sentence",
    "polygon_name",
    "source_url",
    "page_title",
    "search_queries",
    "search_title",
    "search_description",
    "country",
    "world_region",
    "polygon_category",
    "area_size_bin",
    "quality_score",
    "osm_type",
    "osm_id",
    "sentence_id",
    "parse_error",
]

SENTENCE_LABEL_OPTIONS = ["relevant", "irrelevant", "unclear"]


def _clean_cell_text(value) -> str:
    if not isinstance(value, str):
        return ""

    return " ".join(value.split())


def _copy_column_if_present(
    dataframe: pd.DataFrame,
    target_column: str,
    source_column: str,
) -> None:
    if target_column in dataframe.columns:
        dataframe[target_column] = dataframe[target_column].fillna("")
    elif source_column in dataframe.columns:
        dataframe[target_column] = dataframe[source_column].fillna("")
    else:
        dataframe[target_column] = ""


def build_sentence_label_review_dataframe(labeled_df: pd.DataFrame) -> pd.DataFrame:
    review_df = labeled_df.copy().reset_index(drop=True)
    review_df["review_id"] = [
        f"sentence-review-{index:04d}" for index in range(1, len(review_df) + 1)
    ]
    review_df["human_label"] = ""
    review_df["human_notes"] = ""
    _copy_column_if_present(review_df, "source_url", "url")
    _copy_column_if_present(review_df, "page_title", "title")
    review_df["parse_error"] = review_df["parse_error"].fillna("")
    review_df["sentence"] = review_df["sentence"].apply(_clean_cell_text)

    for column in SENTENCE_REVIEW_COLUMNS:
        if column not in review_df.columns:
            review_df[column] = ""

    return review_df[SENTENCE_REVIEW_COLUMNS]


def save_sentence_label_review_xlsx(
    review_df: pd.DataFrame,
    path: str | Path,
) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="review", index=False)

    workbook = load_workbook(path)
    worksheet = workbook["review"]
    worksheet.freeze_panes = "E2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    column_widths = {
        "A": 22,
        "B": 16,
        "C": 30,
        "D": 14,
        "E": 100,
        "F": 30,
        "G": 46,
        "H": 34,
        "I": 52,
        "J": 34,
        "K": 42,
        "L": 20,
        "M": 20,
        "N": 20,
        "O": 14,
        "P": 14,
        "Q": 12,
        "R": 14,
        "S": 28,
        "T": 24,
    }
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet["G"][1:]:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    label_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(SENTENCE_LABEL_OPTIONS)}"',
        allow_blank=True,
    )
    worksheet.add_data_validation(label_validation)
    label_validation.add(f"B2:B{worksheet.max_row}")

    workbook.save(path)
