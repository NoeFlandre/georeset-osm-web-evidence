import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

REVIEW_COLUMNS = [
    "review_id",
    "human_label",
    "human_notes",
    "polygon_name",
    "text_preview",
    "extraction_error",
    "extraction_method",
    "source_url",
    "page_title",
    "search_queries",
    "search_title",
    "search_description",
    "has_wikipedia_articles",
    "text_length",
    "fetch_status",
    "osm_type",
    "osm_id",
]

LABEL_OPTIONS = ["relevant", "irrelevant", "broken", "unclear"]


def clean_text_for_review(text: str | None) -> str:
    if not isinstance(text, str):
        return ""

    return " ".join(text.split())


def make_text_preview(text: str | None, preview_chars: int = 1500) -> str:
    clean_text = clean_text_for_review(text)

    if len(clean_text) <= preview_chars:
        return clean_text

    return clean_text[:preview_chars].rstrip() + "…"


def select_successful_review_rows(
    page_text_df: pd.DataFrame,
    max_rows: int = 30,
    max_rows_per_polygon: int = 2,
) -> pd.DataFrame:
    successful_df = page_text_df[
        page_text_df["fetch_error"].isna()
        & page_text_df["text"].apply(lambda text: isinstance(text, str) and bool(text))
    ].copy()

    successful_df = successful_df.sort_values(
        [
            "has_wikipedia_articles",
            "osm_type",
            "polygon_name",
            "source_url",
        ]
    )

    capped_df = successful_df.groupby("polygon_name", group_keys=False).head(
        max_rows_per_polygon
    )

    return capped_df.head(max_rows).reset_index(drop=True)


def build_human_review_dataframe(
    page_text_df: pd.DataFrame,
    preview_chars: int = 1500,
    max_rows: int = 30,
    max_rows_per_polygon: int = 2,
) -> pd.DataFrame:
    review_df = select_successful_review_rows(
        page_text_df,
        max_rows=max_rows,
        max_rows_per_polygon=max_rows_per_polygon,
    )

    review_df = review_df.sort_values(["polygon_name", "source_url"]).reset_index(
        drop=True
    )
    review_df["review_id"] = [
        f"review-{index:04d}" for index in range(1, len(review_df) + 1)
    ]
    review_df["human_label"] = ""
    review_df["human_notes"] = ""
    review_df["fetch_status"] = review_df["fetch_error"].apply(
        lambda value: "broken" if isinstance(value, str) and value else "fetched"
    )
    review_df["page_title"] = review_df["title"].fillna("")
    review_df["text_preview"] = review_df["text"].apply(
        lambda text: make_text_preview(text, preview_chars=preview_chars)
    )
    review_df["fetch_error"] = review_df["fetch_error"].fillna("")

    for column in ["extraction_method", "extraction_error"]:
        if column not in review_df.columns:
            review_df[column] = ""

        review_df[column] = review_df[column].fillna("")

    return review_df[REVIEW_COLUMNS]


def save_human_review_xlsx(review_df: pd.DataFrame, path: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="review", index=False)

    workbook = load_workbook(path)
    worksheet = workbook["review"]

    worksheet.freeze_panes = "E2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    column_widths = {
        "A": 14,  # review_id
        "B": 16,  # human_label
        "C": 30,  # human_notes
        "D": 28,  # polygon_name
        "E": 90,  # text_preview
        "F": 28,  # extraction_error
        "G": 18,  # extraction_method
        "H": 42,  # source_url
        "I": 34,  # page_title
        "J": 44,  # search_queries
        "K": 34,  # search_title
        "L": 42,  # search_description
        "M": 18,  # has_wikipedia_articles
        "N": 12,  # text_length
        "O": 14,  # fetch_status
        "P": 12,  # osm_type
        "Q": 14,  # osm_id
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for cell in worksheet["H"][1:]:
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    label_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(LABEL_OPTIONS)}"',
        allow_blank=True,
    )
    worksheet.add_data_validation(label_validation)
    label_validation.add(f"B2:B{worksheet.max_row}")

    workbook.save(path)
